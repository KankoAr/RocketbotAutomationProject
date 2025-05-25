import time
import smtplib
import logging 
import configparser
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def procesar_excel():
    archivo_excel = "Base Seguimiento Observ Auditoría al_30042021.xlsx"
    wb = None
    driver = None

    try:
        wb = load_workbook(archivo_excel)
        sheet = wb.active
        
        needs_webdriver = False
        for row_values in sheet.iter_rows(min_row=2, values_only=True):
            if row_values[9] == 'Regularizado':
                needs_webdriver = True
                break
        
        if needs_webdriver:
            try:
                driver = webdriver.Chrome()
                logging.info("WebDriver initialized.")
            except Exception as e:
                logging.error(f"Error initializing WebDriver: {e}", exc_info=True)
                driver = None 
        
        for row_idx, current_row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            proceso_val_for_log = current_row_cells[0].value # For logging in case of error before all values are read
            try:
                proceso = current_row_cells[0].value
                observacion = current_row_cells[1].value
                tipo_riesgo = current_row_cells[2].value
                severidad = current_row_cells[3].value
                plan_accion = current_row_cells[4].value 
                fecha_compromiso = current_row_cells[5].value
                responsable = current_row_cells[6].value
                responsable_area = current_row_cells[7].value 
                correo_responsable = current_row_cells[8].value
                estado = current_row_cells[9].value
                
                if estado == 'Regularizado':
                    if driver:
                        logging.info(f"Processing row {row_idx} for web submission: {proceso}")
                        subir_informacion(driver, proceso, tipo_riesgo, severidad, responsable, fecha_compromiso, observacion)
                    else:
                        logging.warning(f"Skipping web submission for row {row_idx} (proceso {proceso}) as WebDriver is not available.")
                elif estado == 'Atrasado':
                    logging.info(f"Processing row {row_idx} for email: {proceso}")
                    enviar_correo(proceso, estado, observacion, fecha_compromiso, correo_responsable)
            except Exception as e:
                logging.error(f"Error al procesar fila {row_idx} ({proceso_val_for_log if proceso_val_for_log else 'N/A'}): {e}", exc_info=True)
        
    except FileNotFoundError as e:
        logging.error(f"Error: El archivo '{archivo_excel}' no fue encontrado.", exc_info=True)
    except Exception as e:
        logging.error(f"Error en la función procesar_excel antes del bucle de filas: {e}", exc_info=True)
    finally:
        if wb:
            try:
                wb.close()
                logging.info("Excel workbook closed.")
            except Exception as e:
                logging.error(f"Error closing workbook: {e}", exc_info=True)
        if driver:
            try:
                driver.quit()
                logging.info("WebDriver closed.")
            except Exception as e:
                logging.error(f"Error closing WebDriver: {e}", exc_info=True)

def subir_informacion(driver, proceso_original, tipo_riesgo, severidad_original, responsable, fecha_compromiso, observacion):
    # proceso_original and severidad_original are the raw values from Excel for logging purposes.
    # The 'proceso' variable within this function will refer to proceso_original for most logging contexts.
    proceso_for_logging = proceso_original 
    try:
        driver.get("https://roc.myrb.io/s1/forms/M6I8P2PDOZFDBYYG")
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "process")))
        submission_successful = True

        # 1. Process Dropdown
        try:
            # Ensure proceso_original is a string before .lower(), handle if None
            proceso_input_str = str(proceso_original) if proceso_original is not None else ""
            proceso_lower = proceso_input_str.lower()
            
            select_element = driver.find_element(By.ID, 'process')
            select = Select(select_element)
            found_proceso = False
            if proceso_original is not None: # Only attempt to find if there's an input
                for option in select.options:
                    if option.text.lower() == proceso_lower:
                        option.click()
                        found_proceso = True
                        break
            
            if not found_proceso:
                available_options = [opt.text for opt in select.options]
                if proceso_original is None:
                    logging.warning(f"Advertencia (Proceso '{proceso_for_logging}'): Input para 'process' estaba vacío/None. No se seleccionó ninguna opción. Disponibles: {available_options}.")
                else:
                    log_msg = f"Advertencia (Proceso '{proceso_for_logging}'): Opción '{proceso_lower}' (input normalizado desde '{proceso_original}') no encontrada en desplegable 'process'."
                    log_msg += f" Disponibles: {available_options}. Se intentará continuar."
                    logging.warning(log_msg)
                # Not returning here as per previous logic, but this is a critical field.
                # Depending on requirements, one might choose to set submission_successful = False or return.
        except (NoSuchElementException, TimeoutException) as e:
            logging.error(f"Error crítico (Proceso '{proceso_for_logging}'): No se pudo encontrar o interactuar con el desplegable 'process': {e}. Saltando este envío.", exc_info=True)
            return # Critical field, cannot continue this submission

        # 2. Tipo de Riesgo
        try:
            if tipo_riesgo is not None:
                 driver.find_element(By.ID, 'tipo_riesgo').send_keys(str(tipo_riesgo))
            else:
                logging.info(f"Información (Proceso '{proceso_for_logging}'): Campo 'tipo_riesgo' está vacío. Se omitirá.")
        except (NoSuchElementException, ElementNotInteractableException, TimeoutException) as e:
            logging.error(f"Error (Proceso '{proceso_for_logging}'): No se pudo interactuar con campo 'tipo_riesgo': {e}. Continuando con el siguiente campo.", exc_info=True)
            submission_successful = False

        # 3. Severidad Dropdown
        try:
            if severidad_original is not None:
                severidad_input_str = str(severidad_original) # Ensure string before strip
                severidad_stripped = severidad_input_str.strip()
                severidad_dato = driver.find_element(By.ID, 'severidad')
                severidad_select = Select(severidad_dato)
                found_severidad = False
                for option in severidad_select.options:
                    if option.text == severidad_stripped: # Comparison is case-sensitive as per original logic
                        option.click()
                        found_severidad = True
                        break
                if not found_severidad:
                    available_options = [opt.text for opt in severidad_select.options]
                    log_msg = f"Advertencia (Proceso '{proceso_for_logging}'): Opción '{severidad_stripped}' (input normalizado desde '{severidad_original}') no encontrada en desplegable 'severidad'."
                    log_msg += f" Disponibles: {available_options}. Se intentará continuar."
                    logging.warning(log_msg)
            else:
                logging.info(f"Información (Proceso '{proceso_for_logging}'): Campo 'severidad' está vacío. Se omitirá.")
        except (NoSuchElementException, ElementNotInteractableException, TimeoutException) as e:
            logging.error(f"Error (Proceso '{proceso_for_logging}'): No se pudo interactuar con desplegable 'severidad': {e}. Continuando con el siguiente campo.", exc_info=True)
            submission_successful = False

        # 4. Responsable
        try:
            if responsable is not None:
                driver.find_element(By.ID, 'res').send_keys(str(responsable))
            else:
                logging.info(f"Información (Proceso '{proceso_for_logging}'): Campo 'responsable' está vacío. Se omitirá.")
        except (NoSuchElementException, ElementNotInteractableException, TimeoutException) as e:
            logging.error(f"Error (Proceso '{proceso_for_logging}'): No se pudo interactuar con campo 'responsable': {e}. Continuando con el siguiente campo.", exc_info=True)
            submission_successful = False
            
        # 5. Fecha Compromiso
        try:
            if fecha_compromiso is not None and hasattr(fecha_compromiso, 'strftime'):
                fecha_formateada = fecha_compromiso.strftime('%d/%m/%Y')
                driver.find_element(By.ID, 'date').send_keys(fecha_formateada)
            elif fecha_compromiso is not None:
                logging.warning(f"Advertencia (Proceso '{proceso_for_logging}'): fecha_compromiso ('{fecha_compromiso}') no es un objeto de fecha válido. Se enviará como texto: '{str(fecha_compromiso)}'.")
                driver.find_element(By.ID, 'date').send_keys(str(fecha_compromiso))
            else:
                logging.info(f"Información (Proceso '{proceso_for_logging}'): Campo 'fecha_compromiso' está vacío. Se omitirá.")
        except (NoSuchElementException, ElementNotInteractableException, TimeoutException) as e:
            logging.error(f"Error (Proceso '{proceso_for_logging}'): No se pudo interactuar con campo 'fecha_compromiso': {e}. Continuando con el siguiente campo.", exc_info=True)
            submission_successful = False

        # 6. Observacion
        try:
            if observacion is not None:
                driver.find_element(By.ID, 'obs').send_keys(str(observacion))
            else:
                logging.info(f"Información (Proceso '{proceso_for_logging}'): Campo 'observacion' está vacío. Se omitirá.")
        except (NoSuchElementException, ElementNotInteractableException, TimeoutException) as e:
            logging.error(f"Error (Proceso '{proceso_for_logging}'): No se pudo interactuar con campo 'observacion': {e}. Continuando con el siguiente campo.", exc_info=True)
            submission_successful = False

        # Click en submit
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "submit"))).click()
        except (TimeoutException, ElementNotInteractableException) as e:
            logging.error(f"Error crítico (Proceso '{proceso_for_logging}'): No se pudo hacer clic en el botón 'submit': {e}. No se puede confirmar el envío.", exc_info=True)
            submission_successful = False # Mark as not fully successful
            return # Cannot proceed to check confirmation

        if not submission_successful:
            logging.warning(f"Advertencia (Proceso '{proceso_for_logging}'): Se encontraron errores al llenar algunos campos. Es posible que el envío no sea completo o haya fallado.")

        # Verificación de envío de formulario
        try:
            elemento_alerta = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "alert-success")))
            texto_alerta = elemento_alerta.text.strip()

            if "Data sent" in texto_alerta:
                if "Queue ID " in texto_alerta:
                    try:
                        id_cola = texto_alerta.split("Queue ID ")[1].strip()
                        logging.info(f"Información para '{proceso_for_logging}' enviada correctamente. ID de la cola: {id_cola}")
                    except IndexError:
                        logging.warning(f"Información para '{proceso_for_logging}' enviada. 'Queue ID ' encontrado, pero no se pudo extraer el ID. Alerta completa: '{texto_alerta}'")
                else: 
                    logging.warning(f"Información para '{proceso_for_logging}' enviada, pero formato de ID de cola no reconocido. Alerta completa: '{texto_alerta}'")
            else: 
                logging.warning(f"No se pudo confirmar el envío para '{proceso_for_logging}' con 'Data sent'. Alerta: {texto_alerta}")
        except TimeoutException:
            logging.error(f"Error (Proceso '{proceso_for_logging}'): No se encontró la alerta de confirmación después de enviar el formulario.", exc_info=True)
        except Exception as e: 
            logging.error(f"Error inesperado durante la verificación de envío para '{proceso_for_logging}': {e}. Texto de alerta (si disponible): '{texto_alerta if 'texto_alerta' in locals() else 'No disponible'}'", exc_info=True)

    except Exception as e: 
        logging.error(f"Error general al subir la información al formulario para '{proceso_for_logging}': {e}", exc_info=True)

def enviar_correo(proceso, estado, observacion, fecha_compromiso ,correo_responsable):
    config = configparser.ConfigParser()
    try:
        if not config.read('config.ini'):
            logging.error("Error: Archivo 'config.ini' no encontrado. Por favor, crea uno a partir de 'config_example.ini' y completa tus credenciales.")
            return 

        remitente = config.get('SMTP', 'email')
        password_remitente = config.get('SMTP', 'password')

        if not remitente or not password_remitente or remitente == "tu_email@gmail.com" or password_remitente == "tu_password_de_aplicacion":
            logging.error("Error: Credenciales SMTP no configuradas o son las predeterminadas en 'config.ini'. Por favor, actualiza el archivo.")
            return 

        destinatario = correo_responsable
        asunto = f'Estado de proceso: {proceso}'
        fecha_str = fecha_compromiso.strftime('%d/%m/%Y') if hasattr(fecha_compromiso, 'strftime') else str(fecha_compromiso)
        cuerpo = f"Estado del proceso: {estado}\nObservación: {observacion}\nFecha de compromiso: {fecha_str}"
        
        mensaje = MIMEMultipart()
        mensaje['From'] = remitente
        mensaje['To'] = destinatario
        mensaje['Subject'] = asunto
        mensaje.attach(MIMEText(cuerpo, 'plain'))
        
        servidor_smtp = smtplib.SMTP('smtp.gmail.com', 587)
        servidor_smtp.starttls()
        servidor_smtp.login(remitente, password_remitente)
        servidor_smtp.send_message(mensaje)
        servidor_smtp.quit()
        
        logging.info(f"Correo enviado exitosamente a {destinatario} para el proceso '{proceso}'")
    except FileNotFoundError: 
        logging.error("Error: Archivo 'config.ini' no encontrado. Por favor, crea uno a partir de 'config_example.ini' y completa tus credenciales.", exc_info=True)
    except (configparser.NoSectionError, configparser.NoOptionError) as e:
        logging.error(f"Error de configuración en 'config.ini': {e}. Asegúrate de que la sección [SMTP] y las opciones 'email' y 'password' existan.", exc_info=True)
    except smtplib.SMTPAuthenticationError as e:
        logging.error(f"Error de autenticación SMTP para '{proceso}'. Verifica las credenciales en 'config.ini'.", exc_info=True)
    except Exception as e:
        logging.error(f"Error al enviar el correo electrónico para '{proceso}': {e}", exc_info=True)

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(message)s',
        handlers=[logging.StreamHandler()]
    )
    procesar_excel()
    logging.info("El programa ha finalizado la carga de formularios y envio de emails")
